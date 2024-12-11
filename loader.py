# imports  
import numpy as np 
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def load_from_stupid_excel(file_path:str,sheet_name:str)->pd.DataFrame:
    # read file
    items_data=pd.read_excel(file_path,sheet_name=sheet_name,header=3,usecols=[0,1,2,3])
    items_data.columns=["ID"," Name","Product Link","Brand"]

    workbook = openpyxl.load_workbook('Book1.xlsx')
    worksheet = workbook['Sheet1']
    
    start_list=[4,5,6,7,8]
    numbers_data_list=[]

    for _ in range(0,24,1):
        numbers_data=pd.read_excel(file_path,sheet_name=sheet_name,header=3,usecols=start_list)
        week=worksheet[f"{get_column_letter(start_list[1])}3"].value
        
        numbers_data.columns=["Impressions","Clicks","Orders","Items Number","Sales Value"]
        numbers_data["week"]=week
        start_list=[i+5 for i in start_list]
        numbers_data_list.append(numbers_data)
        # print(numbers_data)
        # print(start_list)
        # break
    # print(len(numbers_data_list))

    concat_numbers=pd.concat(numbers_data_list)
    # print(concat_numbers.shape)

    merged_df=pd.merge(items_data,concat_numbers,how="left",left_index=True,right_index=True)    
    # print(merged_df)
    # merged_df.to_csv("merged_df.csv")
    return merged_df

# file_path="Book1.xlsx"    
# sheet_name="Sheet1"

# load_from_stupid_excel(file_path=file_path,sheet_name=sheet_name)